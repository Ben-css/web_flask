from flask import *
import pymongo
from flask_bootstrap import Bootstrap
from flask_moment import Moment
from datetime import datetime,timedelta,timezone
from flask_wtf import FlaskForm
from wtforms import (StringField, BooleanField, DateTimeField,
                     RadioField, SelectField,
                     TextAreaField, SubmitField)
from wtforms.validators import DataRequired
import pytz
import json
from bson import ObjectId
import secrets
import os
import pathlib
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font           
from openpyxl.styles import NamedStyle, PatternFill   


# 連接資料庫
client = pymongo.MongoClient(
    "mongodb+srv://root:root123@cluster0.9cjg4lw.mongodb.net/?retryWrites=true&w=majority")
db = client.test

# 取得目前檔案所在的資料夾 
SRC_PATH =  pathlib.Path(__file__).parent.absolute()
# 結合目前的檔案路徑和static及uploads路徑 
UPLOAD_FOLDER = os.path.join(SRC_PATH,  'static', 'uploads')

def random_filename(filename):
    # 生成隨機的文件名，保留文件的擴展名
    random_name = secrets.token_hex(16)
    _, ext = os.path.splitext(filename)
    return random_name + ext


app = Flask(
    __name__,
    static_folder="static",
    static_url_path="/"
)
bootstrap = Bootstrap(app)
moment = Moment(app)

# 設定session為24小時
app.permanent_session_lifetime = timedelta(hours=24)
# 設定scret_key為亂碼
# app.config['SECRET_KEY'] = os.urandom(24)
# 設定session key
app.secret_key = "Hello World"



# 首頁
@app.route("/")
def index():
    return render_template("index.html")

# 註冊route
@app.route("/signup_page")
def signuppage():
    return render_template("signup.html")

@app.route("/page/forgetPassword")
def forgetPassword_page():
    return render_template("forgetPassword.html")

@app.route("/signup", methods=["POST"])
def signup():
    student_id = request.form['student_id']
    password = request.form['password']
    # email = request.form['email']
    name = request.form['name']
    # number = request.form['number']

    collection = db.users
    result = collection.find_one({
        "student_id": student_id
    })
    if result != None:
        return redirect("/error?msg=學號已被註冊")

    collection.insert_one({
        "student_id": student_id,
        "password": password,
        # "email": email,
        "name": name,
        # "number": number
    })
    return redirect("/")

#登入route
@app.route("/signin", methods=["POST"])
def signin():
    student_id = request.form["student_id"]
    password = request.form["password"]
    # 和資料庫做互動
    collection = db.users
    # 檢查帳號密碼是否正確
    result = collection.find_one({
        "$and": [
            {"student_id": student_id},
            {"password": password},
        ]
    })

    # 登入失敗，到錯誤頁面
    if result == None:
        return redirect("/error?msg=帳號或密碼輸入錯誤")
    # 登入成功，在 session 紀錄會員資訊，到會員頁面
    session["student_id"] = result["student_id"]
    session["name"] = result["name"]
    # session["number"] = result["number"]
    # print(session['student_id'])
    return redirect("/form")

# 會員頁(目前沒用)
# @app.route("/member")
# def member():
#     # form = MyForm()
#     # 如果student_id在session才能登入
#     if "student_id" in session:
#         return render_template("member.html",
#         #    form=form,
#             name=session.get('name'),
#             student_id=session.get('student_id'),         
#             )
#     else:
#         return redirect("/")
    
# 公告 
@app.route("/announcement_list",methods=["GET"])
def announcement_list():
    # form = MyForm()
    if "student_id" in session:
        collection = db.announcements

        query = {}
        query["status"] = {"$in": ['上架']}

        announcements = list(collection.find(query).sort("created_at", pymongo.DESCENDING))
        
        for annc in announcements:
            annc["_id"] = str(annc["_id"])
        
        return render_template("announcement_list.html",
            name=session.get('name'),
            student_id=session.get('student_id'),
            announcements = announcements,
            )
    else:
        return redirect("/error?msg=未進行登入，請先登入")    


@app.route("/announcement/<_id>",methods=["GET"])
def announcement(_id):
    # form = MyForm()
    if "student_id" in session:
        # student_id=session.get('student_id')
        collection = db.announcements
        _id = ObjectId(_id)
        announcement = collection.find_one({"_id": _id})
        # print(bool(collection.find_one({"_id": _id})))
        # print(type(_id))
        # for student in form_list:
        #     student["_id"] = str(student["_id"])

        # print(form_list)
        return render_template("/announcement.html",
        #    form=form,
            name=session.get('name'),
            student_id=session.get('student_id'),
            announcement = announcement,
            )
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 表單頁
@app.route("/form",methods=["GET"])
def form():
    # form = MyForm()
    # 如果student_id在session才能登入
    if "student_id" in session:
        student_id=session.get('student_id')
        collection = db.forms
        announcements = db.announcements
        form_list =  list(collection.find({"student_id": student_id}).sort("submit_at", pymongo.DESCENDING))
        
        for student in form_list:
            student["_id"] = str(student["_id"])

        # 將_id轉成str
        announcements_list = announcements.find({"status": '上架'}).sort("created_at", pymongo.DESCENDING)
        # for anc in announcements_list:
        #     anc["_id"] = str(anc["_id"])
        # with open('static/fix_items.json', 'r', encoding='UTF-8') as f:
        #     fix_items_data = json.load(f)

        # print(form_list)
        # print(fix_items_data)
        return render_template("form.html",
        #    form=form,
            name=session.get('name'),
            student_id=session.get('student_id'),
            form_list=form_list,
            announcements = announcements_list
            )
    else:
        return redirect("/error?msg=未進行登入，請先登入")


# 進度表
@app.route("/tesk_info/<form_id>")
def get_progress(form_id):
    
    student_id=session.get('student_id')

    collection = db.forms
    # 列出該位使用者提交的forms
    form_list =  list(collection.find({"student_id": student_id}).sort("submit_at", pymongo.DESCENDING))
    # 尋找特定的forms，以顯示詳細資訊
    form = collection.find_one({"_id": ObjectId(form_id), "student_id": student_id})
    # form = collection.find_one({"_id": form_id, "student_id": student_id})
    # print(bool(collection.find_one({"_id": form_id})))
    # print(type(collection["_id"]))
    # print(form)
    # for student in form_list:
    #     student["_id"] = str(student["_id"])

    # footer_img = '/img/dorm_logo_black.png'

    if form:
        return render_template("tesk_info.html",
                                form_list=form_list,
                                name=session.get('name'),
                                student_id=session.get('student_id'),
                                form=form,                      
                        )
    else:
        # 如果沒找到相應的資料，可以處理錯誤或重定向到其他頁面
        return "找不到相應的資料", 404


# 表單驗證、儲存
@app.route('/submit_form',methods=["POST"])
def submit_form():

    twtime = pytz.timezone('Asia/Taipei')
    # submit_at = datetime.today()
    # print(submit_at)
    
    # 設定utc時區
    utc_now = datetime.utcnow()
    # 設定8小時
    taipei_offset = timedelta(hours=8)

    #加8小時
    submit_at_taipei = utc_now.replace(tzinfo=timezone.utc) + taipei_offset

    # print(submit_at_taipei)
    # submit_at_str = submit_at.strftime("%Y.%m.%d %H:%M:%S")
    random_name = ""  # 默認值

    file = request.files['image']
    if file.filename != '':
        random_name = random_filename(file.filename)
        file.save(os.path.join(UPLOAD_FOLDER, random_name))
        # print(type(random_name))
        # return 'File uploaded as: ' + random_name
        # file.save(os.path.join(UPLOAD_FOLDER, file.filename))
    # print(type(submit_at))
    # result = collection.find_one({
    #         "$and": [
    #             {"student_id": student_id},
    #             {"password": password},
    #         ]
    #     })

    #     # 登入失敗，到錯誤頁面
    #     if result == None:
    if request.method == "POST":
        student_id = session['student_id']
        name = session['name']
        # number = session['number']
        dorms = request.form["dorms"]
        location = request.form["location"]
        location_detail = request.form["location_detail"]
        fix_items = request.form["fix_items"]
        other_fix_items = request.form.get("other_fix_items",'')
        fix_explain = request.form["fix_explain"]
        # print(request.form.get("other_fix_items"))
        # 進度說明預設為空
        progress_explain = ""

        forms = db.forms
        forms.insert_one({
            "student_id": student_id,
            "name": name,
            "dorms": dorms,
            # "number": number,
            "location": location,
            "location_detail":location_detail,
            "fix_items": fix_items,
            "other_fix_items": other_fix_items,
            "fix_explain": fix_explain,
            "status": "待確認",
            "submit_at": submit_at_taipei,
            "progress_explain": progress_explain,
            "image": random_name,
            "update_at": submit_at_taipei,
            "fixing_update_at": '',
            "finish_update_at": '',
        })

    # 刪除資料用
    # collection = db.forms
    # result = collection.delete_many({"name": {"$exists": True}})
    #     result = collection.delete_many({
    #         "name": "林東科"
    # })
    #     print("實際上被刪除的資料",str(result.deleted_count) + "筆")

    # 記得重新導向回/member !!
    # 不然用render_template("/member",form=form)會被導到/forms
    return redirect('/form')
# redirect(url_for('member'))


# 管理員登入
@app.route("/manager_signin_page", methods=["GET"])
def manager_signin_page():
    return render_template('manager_login.html')


@app.route("/manager_signin", methods=["POST"])
def manager_signin():
    account = request.form["account"]
    password = request.form["password"]
    # 和資料庫做互動
    collection = db.managers
    # 檢查帳號密碼是否正確
    result = collection.find_one({
        "$and": [
            {"account": account},
            {"password": password},
        ]
    })

    # # 登入失敗，到錯誤頁面
    if result == None:
        return redirect("/error?msg=帳號或密碼輸入錯誤")
    # 登入成功，在 session 紀錄會員資訊，到會員頁面
    session["account"] = result["account"]
    session["manager_name"] = result["manager_name"]
    # print(session['account'])
    return redirect("/manager/page")

@app.route("/manager/page",methods=['GET'])
def manager_page():
    # personal_info = db.managers
    if "manager_name" in session:
        collection = db.forms    
        # form_list =  list(collection.find().sort("submit_at", pymongo.DESCENDING))

        # form = collection.find_one({"_id": ObjectId(form_id), "student_id": student_id})
            
        dorm = request.args.getlist("dorm")
        status = request.args.getlist("status")
        start_at = request.args.get("start_at",'')
        end_at = request.args.get("end_at",'')

        # 創建初始的查詢條件，這裡假設 dorm 和 status 都是列表
        query = {}

        # 添加 dorm 的篩選條件
        if "all_dorms" in dorm or dorm == []:
            # 不需要進一步處理其他宿舍的條件
            pass
        else:
            query["dorms"] = {"$in": dorm}

        # 添加 status 的篩選條件
        if "all_status" in status or status == []:
            # 不需要進一步處理其他狀態的條件
            pass
        else:
            query["status"] = {"$in": status}

        # 添加 start_at 和 end_at 的篩選條件
        if start_at != '' and end_at != '':
            start_date = datetime.strptime(start_at, "%Y-%m-%d")
            end_date = datetime.strptime(end_at, "%Y-%m-%d")
            # 將開始時間設成0:00，結束時間設成23:59
            start_date = start_date.replace(hour=0, minute=0, second=0)
            end_date = end_date.replace(hour=23, minute=59, second=59)
            query["submit_at"] = {"$gte": start_date, "$lte": end_date}

        # 查詢並排序
        form_list = list(collection.find(query).sort("submit_at", pymongo.DESCENDING))

        for student in form_list:
            student["_id"] = str(student["_id"])

        return render_template('manager_page.html',
            form_list =  form_list,
            dorm = dorm,
            status = status,
            start_at = start_at,
            end_at = end_at
            )
    else:
        return redirect("/error?msg=未進行登入，請登入")

@app.route("/manager/excel_export", methods=['POST'])
def manager_excel_export():
    try:
        data = request.get_json()
        collection = db.forms
        id_values = [ObjectId(value) for value in data.get('idValues', [])]

        # 使用 sort 方法對查询结果進行排序
        excel_list = list(collection.find({"_id": {"$in": id_values}}).sort("submit_at", pymongo.DESCENDING))
        df = pd.DataFrame(excel_list)

        # 映射英文列名和中文列名的字典
        column_mapping = {
            'student_id': '學號',
            'name': '學生姓名',
            'dorms': '棟別',
            'location': '地點',
            'location_detail': '詳細地點',
            'fix_items': '維修項目',
            'other_fix_items': '其他維修項目',
            'fix_explain': '維修說明',
            'submit_at': '繳交時間',
            'image': '損壞圖片'
        }

        # 提取特定列的數据
        selected_columns = list(column_mapping.keys())
        df_selected = pd.DataFrame(df, columns=selected_columns)

        # 重命名列名
        df_selected.rename(columns=column_mapping, inplace=True)

        # 生成當前时间的字符串，用作 Excel 文件名
        current_time_str = datetime.now().strftime("%Y%m%d%H%M%S")
        excel_filename = f"excel_output_{current_time_str}.xlsx"

        # 指定資料夾
        folder_path = 'download_excel'
        os.makedirs(folder_path, exist_ok=True)  # 如果文件夹不存在，則創建文件夹

        # 提供完整的路徑以保存 Excel 文件
        full_path = os.path.join(folder_path, excel_filename)

        # 使用 ExcelWriter 設置樣式
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            df_selected.to_excel(writer, index=True, sheet_name='Sheet1')

            # 設定sheet 
            sheet = writer.sheets['Sheet1']

            # 將標題設為 淡藍色
            # header_style = openpyxl.styles.NamedStyle(name='header_style', fill=openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'))
            # for cell in sheet["1:1"]:
            #     cell.style = header_style

            # 設置欄寬
            sheet.column_dimensions['A'].width = 5
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 10
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
            sheet.column_dimensions['H'].width = 15
            sheet.column_dimensions['I'].width = 25
            sheet.column_dimensions['J'].width = 20
            sheet.column_dimensions['K'].width = 40

            # 設置行高
            for row_num in range(1, sheet.max_row + 1):
                sheet.row_dimensions[row_num].height = 25

            # 設置標題樣式
            title_style = openpyxl.styles.NamedStyle(name='header_style', fill=openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'))
            title_style.font = openpyxl.styles.Font(bold=True)
            title_style.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            title_style.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))

            # 应用标题样式到每个单元格
            for cell in sheet["1:1"]:
                cell.style = title_style

        # 返回重定向到下载链接的响应
        download_link = url_for('download_excel', filename=excel_filename)
        response = {'message': 'Data received successfully', 'excel_download_link': download_link}
        # print(download_link)

        # 直接重定向到下载链接
        return redirect(download_link)
    except Exception as e:
        print(str(e))
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/manager/download_excel/<filename>', methods=['GET'])
def download_excel(filename):
    folder_path = 'download_excel'
    full_path = os.path.join(folder_path, filename)
    # return redirect(filename)
    return send_file(full_path, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', download_name=filename)

# 管理者的維修單編輯
@app.route("/manager/tesk_info/<_id>",methods=['GET'])
def manager_tesk_info(_id):
    if "manager_name" in session:
        collection = db.forms
        _id = ObjectId(_id)
        form_info = collection.find_one({"_id": _id})
        # print(bool(collection.find_one({"_id": _id})))
        # print(type(_id))
        # for student in form_list:
        form_info["_id"] = str(form_info["_id"])

        # print(form_list)
        return render_template("/manager_tesk_info.html",
        #    form=form,
            # name=session.get('name'),
            # student_id=session.get('student_id'),
            form_info = form_info,
            )
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者的維修表單更新
@app.route("/manager/tesk_update/<_id>",methods=['POST'])
def manager_tesk_update(_id):
    if "manager_name" in session:
        _id = ObjectId(_id)

        # 進度說明預設為空
        status = request.form["status"]
        progress_explain = request.form["progress_explain"]
        manager_name = session['manager_name']

        result = db.forms.find_one({"_id": _id})

        # 找到這項
        if status == '處理中' or status == '已完成':
            finish_update_at_value = ''
            fixing_update_at_value = ''

            # Check if the document exists and contains the finish_update_at field
            if result and "fixing_update_at" in result:
                fixing_update_at_value = result["fixing_update_at"]
            if result and "finish_update_at" in result:
                finish_update_at_value = result["finish_update_at"]

            if status == '處理中':
                fixing_update_at = datetime.today()
                finish_update_at = finish_update_at_value
            elif status == '已完成':
                finish_update_at = datetime.today()
                fixing_update_at = fixing_update_at_value

            update_data = {
                "$set": {
                    "manager_name": manager_name,
                    "status": status,
                    "update_at": datetime.today(),
                    "progress_explain": progress_explain,
                    "fixing_update_at": fixing_update_at,
                    "finish_update_at": finish_update_at,
                }
            }
        else:
            # 如果 status 不是 '處理中' 或 '已完成'，則不修改 fixing_update_at 和 finish_update_at
            update_data = {
                "$set": {
                    "manager_name": manager_name,
                    "status": status,
                    "update_at": datetime.today(),
                    "progress_explain": progress_explain,
                }
            }

        forms = db.forms
        forms.update_one({"_id": _id}, update_data)

        redirect_url = url_for('manager_tesk_info', _id=_id)
        return redirect(redirect_url)
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者的公告管理
@app.route("/manager/announcement",methods=['GET'])
def manager_announcement():
    if "manager_name" in session:
        collection = db.announcements    
        keyword = request.args.get("keyword", '')
        # print(keyword)
        # 創建初始的查詢條件，這裡假設 dorm 和 status 都是列表
        query = {}
        if keyword != '':
            regex_pattern = re.compile(f".*{re.escape(keyword)}.*", re.IGNORECASE)
            query = {
                "$or": [
                    {"creator": {"$regex": regex_pattern}},
                    {"title": {"$regex": regex_pattern}},
                ]
            }
            # print(query) 
            # announcement_list = list(collection.find(query).sort("created_at", pymongo.DESCENDING))
        # else:
            # announcement_list =  list(collection.find().sort("created_at", pymongo.DESCENDING))
        
        status = request.args.getlist("status")
        print(status)
        # 添加 status 的篩選條件
        if "all_status" in status or status == []:
            # 不需要進一步處理其他狀態的條件
            pass
        else:
            query["status"] = {"$in": status}

        announcement_list = list(collection.find(query).sort("created_at", pymongo.DESCENDING))
        
        # print(anncouncement_list)
        # form = collection.find_one({"_id": ObjectId(form_id), "student_id": student_id})

        # other_fix_items = request.form.get("other_fix_items",'')

        for annc in announcement_list:
            annc["_id"] = str(annc["_id"])

        return render_template('manager_announcement.html',
                announcement_list = announcement_list,
                keyword = keyword,
                status = status               
                )
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者新增公告頁
@app.route("/manager/announcement/add",methods=['GET'])
def manager_announcement_add():
    if "manager_name" in session:
        # collection = db.announcements    
        # announcement_list =  list(collection.find().sort("created_at", pymongo.DESCENDING))
        # print(anncouncement_list)
        # form = collection.find_one({"_id": ObjectId(form_id), "student_id": student_id})
            
        # for annc in announcement_list:
        #     annc["_id"] = str(annc["_id"])
        current_url = request.url

        # 檢查 URL 是否包含 "add" 或 "edit"
        if "add" in current_url:
            title = "新增文章"
        elif "edit" in current_url:
            title = "編輯文章"

        action = "/manager/announcement/add/submit"

        return render_template('manager_announcement_edit.html',
                title = title,
                action = action)
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者新增公告_POST
@app.route("/manager/announcement/add/submit",methods=['POST'])
def manager_announcement_add_submit():
    if "manager_name" in session:
        twtime = pytz.timezone('Asia/Taipei')
        # created_at = datetime.today()
        # print(created_at)

        # 設定utc時區
        utc_now = datetime.utcnow()
        # 設定8小時
        taipei_offset = timedelta(hours=8)

        #加8小時
        submit_at_taipei = utc_now.replace(tzinfo=timezone.utc) + taipei_offset
        
        creator = session['manager_name']
        title = request.form["title"]
        content = request.form["content"]
        status = request.form["status"]
        is_top = request.form.get('is_top', 'off')

        announcements = db.announcements
        announcements.insert_one({
            "title": title,
            "content": content,
            "creator": creator,
            "status": status,
            "is_top": is_top,
            "created_at": submit_at_taipei,
            "updated_at": submit_at_taipei,
        })
        # 刪除資料用
        #     collection = db.announcements
        #     result = collection.delete_many({
        #         "name": "柏安"
        # })

        return redirect("/manager/announcement")
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者的公告編輯
@app.route("/manager/announcement/edit/<_id>",methods=['GET'])
def manager_announcement_edit(_id):
    if "manager_name" in session:
        collection = db.announcements
        _id = ObjectId(_id)
        announcement_info = collection.find_one({"_id": _id})
        # print(bool(collection.find_one({"_id": _id})))
        # print(type(_id))
        # for student in form_list:
        announcement_info["_id"] = str(announcement_info["_id"])

        current_url = request.url
        # 檢查 URL 是否包含 "add" 或 "edit"
        if "add" in current_url:
            title = "新增文章"
        elif "edit" in current_url:
            title = "編輯文章"
        
        action = "/manager/announcement/edit/"+str(_id)+"/post"

        return render_template('manager_announcement_edit.html',
                announcement_info = announcement_info,
                title = title,
                action = action)
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者的公告編輯_POST
@app.route("/manager/announcement/edit/<_id>/post",methods=['POST'])
def manager_announcement_edit_post(_id):
    if "manager_name" in session:
        _id = ObjectId(_id)
        filter = {"_id": _id}
        announcement = db.announcements

        twtime = pytz.timezone('Asia/Taipei')
        updated_at = datetime.now(twtime)

        # title,content,status,is_top = ''
        # creator = session['name']
        title = request.form["title"]
        content = request.form["content"]
        status = request.form["status"]
        is_top = request.form.get('is_top', 'off')

        update_data = {
        "$set": {
            "title": title,
            "content": content,
            # "creator": creator,
            "status": status,
            "is_top": is_top,
            # "created_at": created_at,
            "updated_at": updated_at,
            }
        }
        # 用_id塞選後將資料放入
        
        announcement.update_one(filter, update_data)

        return redirect("/manager/announcement")
    else:
        return redirect("/error?msg=未進行登入，請先登入")

# 管理者驗證
@app.route("/admin", methods=["GET"])
def admin():
    return render_template("admin.html")

# 管理員註冊頁
@app.route("/manager_signup_page", methods=["POST"])
def manager_signup_page():
    verification = request.form['verification-code']
    if verification == "99999999":
        session["verification-code"] = verification
        return render_template('manager_signup.html')
    else:
        return redirect("/admin?msg=驗證碼輸入錯誤")

# 管理員註冊
@app.route("/manager_signup", methods=["POST"])
def manager_signup():
    collection = db.managers
    manager_name = request.form['manager_name']
    account = request.form['account']
    password = request.form['password']

    result = collection.find_one({
        "$or": [
            {"manager_name": manager_name},
            {"account": account},
        ]
    })
    if result != None:
        return redirect("/error?msg=該管理員帳號已被註冊")

    collection.insert_one({
        "manager_name": manager_name,
        "account": account,
        "password": password,
    })
    return redirect("/")
    # return redirect("/manager_signup_page")

# 錯誤route
@app.route("/error")
def error():
    message = request.args.get("msg", "發生錯誤")
    return render_template("error.html", message=message)


# 使用者登出
@app.route("/signout")
def signout():
    del session['name']
    del session['student_id']
    # del session["number"]
    return redirect("/")

@app.route("/manager/signout")
def manager_signout():
    del session['account']
    del session['manager_name']
    return redirect("/")

app.run(
    host= '0.0.0.0', #任何ip都可訪問
    port=8001, 
    debug=True
    )
